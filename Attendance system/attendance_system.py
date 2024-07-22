# import face_recognition
# import cv2
# import os
# import zipfile
# from openpyxl import Workbook, load_workbook
# from datetime import datetime
# from PIL import Image
# import numpy as np

# # Load the known faces and names
# known_face_encodings = []
# known_face_names = []

# dataset_path = "/Users/shauryan/Documents/UWINDSOR/SEM 2/ADT/ADT Project/CnCAP/Attendance system/dataset"
# for filename in os.listdir(dataset_path):
#     if filename.endswith(".jpg"):
#         image_path = os.path.join(dataset_path, filename)
#         try:
#             # Open the image file
#             with Image.open(image_path) as img:
#                 # Convert image to RGB format if it's not already in that format
#                 rgb_image = img.convert("RGB")
#                 # Convert the PIL image to a numpy array
#                 image = np.array(rgb_image)
#                 if image.ndim == 3 and image.shape[2] == 3:  # Ensure the image is 8-bit RGB
#                     # Process the image with face_recognition
#                     face_encodings = face_recognition.face_encodings(image)
#                     if face_encodings:  # Check if face encodings are found
#                         face_encoding = face_encodings[0]
#                         known_face_encodings.append(face_encoding)
#                         known_face_names.append(os.path.splitext(filename)[0])
#                     else:
#                         print(f"No face encodings found in image: {filename}")
#                 else:
#                     print(f"Image {filename} is not in 8-bit RGB format")
#         except Exception as e:
#             print(f"Error processing file {filename}: {e}")

# # Initialize variables
# face_locations = []
# face_encodings = []
# face_names = []

# # Create or load the Excel workbook
# try:
#     if os.path.exists("attendance.xlsx"):
#         wb = load_workbook("attendance.xlsx")
#         ws = wb.active
#     else:
#         raise FileNotFoundError
# except (FileNotFoundError, zipfile.BadZipFile):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Attendance"
#     ws.append(["Name", "Date", "Time"])

# # Function to log attendance
# def log_attendance(name):
#     now = datetime.now()
#     date = now.strftime("%Y-%m-%d")
#     time = now.strftime("%H:%M:%S")
#     ws.append([name, date, time])
#     wb.save("attendance.xlsx")

# # Open a video file
# video_path = "/Users/shauryan/Documents/UWINDSOR/SEM 2/ADT/ADT Project/CnCAP/Attendance system/video.mp4"  # Change this to your video file path
# video_capture = cv2.VideoCapture(video_path)

# while video_capture.isOpened():
#     ret, frame = video_capture.read()
#     if not ret:
#         break
    
#     # Convert the frame to RGB format
#     rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    
#     # Resize frame for faster processing
#     small_frame = cv2.resize(rgb_frame, (0, 0), fx=0.25, fy=0.25)
    
#     # Find all the faces and face encodings in the current frame
#     face_locations = face_recognition.face_locations(small_frame)
#     face_encodings = face_recognition.face_encodings(small_frame, face_locations)
    
#     face_names = []
#     for face_encoding in face_encodings:
#         matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
#         name = "Unknown"
        
#         face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
#         best_match_index = min(range(len(face_distances)), key=face_distances.__getitem__)
#         if matches[best_match_index]:
#             name = known_face_names[best_match_index]
        
#         face_names.append(name)
#         log_attendance(name)
    
#     # Display the results
#     for (top, right, bottom, left), name in zip(face_locations, face_names):
#         top *= 4
#         right *= 4
#         bottom *= 4
#         left *= 4
        
#         cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
#         cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
#         font = cv2.FONT_HERSHEY_DUPLEX
#         cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
    
#     # Display the resulting frame
#     cv2.imshow('Video', frame)
    
#     if cv2.waitKey(1) & 0xFF == ord('q'):
#         break

# video_capture.release()
# cv2.destroyAllWindows()

import face_recognition
import cv2
import os
import zipfile
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image
import numpy as np

# Load the known faces and names
known_face_encodings = []
known_face_names = []

dataset_path = "dataset"
for filename in os.listdir(dataset_path):
    if filename.endswith(".jpg"):
        image_path = os.path.join(dataset_path, filename)
        try:
            # Open the image file
            with Image.open(image_path) as img:
                # Convert image to RGB format if it's not already in that format
                rgb_image = img.convert("RGB")
                # Convert the PIL image to a numpy array
                image = np.array(rgb_image)
                if image.ndim == 3 and image.shape[2] == 3:  # Ensure the image is 8-bit RGB
                    # Process the image with face_recognition
                    face_encodings = face_recognition.face_encodings(image)
                    if face_encodings:  # Check if face encodings are found
                        face_encoding = face_encodings[0]
                        known_face_encodings.append(face_encoding)
                        known_face_names.append(os.path.splitext(filename)[0])
                    else:
                        print(f"No face encodings found in image: {filename}")
                else:
                    print(f"Image {filename} is not in 8-bit RGB format")
        except Exception as e:
            print(f"Error processing file {filename}: {e}")

# Initialize variables
face_locations = []
face_encodings = []
face_names = []

# Create or load the Excel workbook
try:
    if os.path.exists("attendance.xlsx"):
        wb = load_workbook("attendance.xlsx")
        ws = wb.active
    else:
        raise FileNotFoundError
except (FileNotFoundError, zipfile.BadZipFile):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time"])

# Set to keep track of logged names
logged_names = set()

# Function to log attendance
def log_attendance(name):
    if name not in logged_names:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        ws.append([name, date, time])
        wb.save("attendance.xlsx")
        logged_names.add(name)

# Open a video file
video_path = "/Users/shauryan/Documents/UWINDSOR/SEM 2/ADT/ADT Project/CnCAP/Attendance system/video.mp4"  # Change this to your video file path
video_capture = cv2.VideoCapture(video_path)

while video_capture.isOpened():
    ret, frame = video_capture.read()
    if not ret:
        break
    
    # Convert the frame to RGB format
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    
    # Resize frame for faster processing
    small_frame = cv2.resize(rgb_frame, (0, 0), fx=0.25, fy=0.25)
    
    # Find all the faces and face encodings in the current frame
    face_locations = face_recognition.face_locations(small_frame)
    face_encodings = face_recognition.face_encodings(small_frame, face_locations)
    
    face_names = []
    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
        
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = min(range(len(face_distances)), key=face_distances.__getitem__)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]
        
        face_names.append(name)
        log_attendance(name)
    
    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4
        
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
    
    # Display the resulting frame
    cv2.imshow('Video', frame)
    
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
